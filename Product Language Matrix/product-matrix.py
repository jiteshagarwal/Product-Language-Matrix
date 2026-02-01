import random
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# --- CONFIGURATION ---
filename = "Smart_Availability_Matrix.xlsx"
regions = ["USA", "Europe", "UAE", "APAC", "Saudi"]
languages = ["English", "Hindi", "Spanish", "German", "Arabic", "French", "Italian"]
services = [
    "ASR",
    "Redaction",
    "Conversation Facts",
    "Gen AI Disposition",
    "ITN",
    "Intent (UniFit)",
    "Entity (NER/Spacy)",
    "Gen AI Summary",
    "Pro-active Service",
    "SSA-LLM",
    "KaaS"
]
products = ["SSA", "RTGA", "CIA", "CRA"]

# Service statuses (for Service Dashboard)
service_statuses = ["General Availability", "Limited Availability", "Not Supported"]

# Product statuses (for Product Dashboard)
product_statuses = [
    "Full Support (In-Region)",
    "Full Support (Cross-Region)",
    "Limited Availability",
    "Not Supported"
]

# Dependency Logic (For Reference/Comments)
# SSA: ASR, SSA-LLM, KaaS
# RTGA: ASR, Intent, Entity, Gen AI Summary, Gen AI Disposition, Redaction, KaaS, Pro-active Service
# CIA: ASR, Gen AI Disposition, Redaction, Conversation Facts
# CRA: ASR, Redaction

wb = Workbook()

# ==========================================
# SHEET 1: SERVICE INPUT (The Source of Truth)
# ==========================================
ws_data = wb.active
ws_data.title = "Service_Input"
headers = ["Region", "Service", "Language", "Status", "Lookup_Key"]
ws_data.append(headers)

# Formatting
header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws_data[1]:
    cell.fill = header_fill
    cell.font = header_font

# Generate Data (Full Factorial for SERVICES only)
row_num = 2
for r in regions:
    for s in services:
        for l in languages:
            # Random Status
            status = random.choices(service_statuses, weights=[45, 30, 25], k=1)[0]
            # Key: Region|Service|Language
            key_formula = f'=A{row_num}&"|"&B{row_num}&"|"&C{row_num}'
            ws_data.append([r, s, l, status, key_formula])
            row_num += 1

ws_data.column_dimensions['E'].hidden = True

# ==========================================
# STYLING HELPERS - Service Dashboard (3 statuses)
# ==========================================
def apply_service_formatting(ws, start_row, start_col, num_cols, num_rows):
    ws.sheet_view.showGridLines = False
    matrix_range = f"{get_column_letter(start_col)}{start_row+1}:{get_column_letter(start_col+num_cols-1)}{start_row+num_rows}"

    # Green (General Availability)
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    dxf_green = DifferentialStyle(fill=green_fill, font=Font(color="FFFFFF"))
    rule_green = Rule(type="containsText", operator="containsText", text="General Availability", dxf=dxf_green)
    rule_green.formula = [f'NOT(ISERROR(SEARCH("General Availability", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_green)

    # Yellow (Limited Availability)
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    dxf_yellow = DifferentialStyle(fill=yellow_fill)
    rule_yellow = Rule(type="containsText", operator="containsText", text="Limited Availability", dxf=dxf_yellow)
    rule_yellow.formula = [f'NOT(ISERROR(SEARCH("Limited Availability", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_yellow)

    # Red (Not Supported)
    red_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    dxf_red = DifferentialStyle(fill=red_fill, font=Font(color="FFFFFF"))
    rule_red = Rule(type="containsText", operator="containsText", text="Not Supported", dxf=dxf_red)
    rule_red.formula = [f'NOT(ISERROR(SEARCH("Not Supported", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_red)

# ==========================================
# STYLING HELPERS - Product Dashboard (4 statuses)
# ==========================================
def apply_product_formatting(ws, start_row, start_col, num_cols, num_rows):
    ws.sheet_view.showGridLines = False
    matrix_range = f"{get_column_letter(start_col)}{start_row+1}:{get_column_letter(start_col+num_cols-1)}{start_row+num_rows}"

    # Dark Green (Full Support - In-Region)
    dark_green_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    dxf_dark_green = DifferentialStyle(fill=dark_green_fill, font=Font(color="FFFFFF"))
    rule_dark_green = Rule(type="containsText", operator="containsText", text="Full Support (In-Region)", dxf=dxf_dark_green)
    rule_dark_green.formula = [f'NOT(ISERROR(SEARCH("Full Support (In-Region)", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_dark_green)

    # Light Green (Full Support - Cross-Region)
    light_green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    dxf_light_green = DifferentialStyle(fill=light_green_fill, font=Font(color="FFFFFF"))
    rule_light_green = Rule(type="containsText", operator="containsText", text="Full Support (Cross-Region)", dxf=dxf_light_green)
    rule_light_green.formula = [f'NOT(ISERROR(SEARCH("Full Support (Cross-Region)", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_light_green)

    # Yellow (Limited Availability)
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    dxf_yellow = DifferentialStyle(fill=yellow_fill)
    rule_yellow = Rule(type="containsText", operator="containsText", text="Limited Availability", dxf=dxf_yellow)
    rule_yellow.formula = [f'NOT(ISERROR(SEARCH("Limited Availability", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_yellow)

    # Red (Not Supported)
    red_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    dxf_red = DifferentialStyle(fill=red_fill, font=Font(color="FFFFFF"))
    rule_red = Rule(type="containsText", operator="containsText", text="Not Supported", dxf=dxf_red)
    rule_red.formula = [f'NOT(ISERROR(SEARCH("Not Supported", {matrix_range})))']
    ws.conditional_formatting.add(matrix_range, rule_red)

# ==========================================
# SHEET 2: SERVICE DASHBOARD
# ==========================================
ws_serv = wb.create_sheet("Service_Dashboard")

# Header Controls
ws_serv['B2'] = "SERVICE AVAILABILITY MATRIX"
ws_serv['B2'].font = Font(size=16, bold=True, color="203764")
ws_serv['B4'] = "Select Region:"
ws_serv['C4'] = regions[0]  # Default
ws_serv['C4'].font = Font(bold=True)
ws_serv['C4'].border = Border(bottom=Side(style='thin'))

# Dropdown
dv = DataValidation(type="list", formula1=f'"{",".join(regions)}"', allow_blank=False)
ws_serv.add_data_validation(dv)
dv.add(ws_serv['C4'])

# Grid Setup
start_row = 7
start_col = 2
ws_serv.cell(row=start_row, column=1, value="Language").font = Font(bold=True)

# Headers (Services)
for i, s in enumerate(services):
    c = start_col + i
    cell = ws_serv.cell(row=start_row, column=c, value=s)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2F5597", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws_serv.column_dimensions[get_column_letter(c)].width = 20

# Rows (Languages) & Logic
for i, lang in enumerate(languages):
    r = start_row + 1 + i
    ws_serv.cell(row=r, column=1, value=lang).font = Font(bold=True)

    for j, s in enumerate(services):
        c = start_col + j
        # INDEX MATCH to look up Service Status directly
        formula = f'=INDEX(Service_Input!$D:$D, MATCH($C$4&"|"&{get_column_letter(c)}${start_row}&"|"&$A{r}, Service_Input!$E:$E, 0))'
        cell = ws_serv.cell(row=r, column=c, value=formula)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

apply_service_formatting(ws_serv, start_row, start_col, len(services), len(languages))

# ==========================================
# SHEET 3: PRODUCT DASHBOARD (The Smart One)
# ==========================================
ws_prod = wb.create_sheet("Product_Dashboard")

# Header Controls
ws_prod['B2'] = "PRODUCT AVAILABILITY MATRIX (Auto-Calculated)"
ws_prod['B2'].font = Font(size=16, bold=True, color="203764")
ws_prod['B4'] = "Select Region:"
ws_prod['C4'] = regions[0]
ws_prod['C4'].font = Font(bold=True)
ws_prod['C4'].border = Border(bottom=Side(style='thin'))

# Dropdown
ws_prod.add_data_validation(dv)  # Use same validation
dv.add(ws_prod['C4'])

# Legend Note
ws_prod['E4'] = "*Cross-Region = service not available locally but available in another region"
ws_prod['E4'].font = Font(italic=True, size=9, color="555555")

# Grid Setup
start_row = 7
start_col = 2
ws_prod.cell(row=start_row, column=1, value="Language").font = Font(bold=True)

# Headers (Products)
for i, p in enumerate(products):
    c = start_col + i
    cell = ws_prod.cell(row=start_row, column=c, value=p)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="203764", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws_prod.column_dimensions[get_column_letter(c)].width = 25

# Total number of regions (for Cross-Region check)
num_regions = len(regions)

# HELPER FUNCTIONS FOR FORMULA GENERATION
def get_local_status(service_name, row_idx):
    """Get status of service in selected region"""
    return f'INDEX(Service_Input!$D:$D, MATCH($C$4&"|{service_name}|"&$A{row_idx}, Service_Input!$E:$E, 0))'

def get_ga_count_globally(service_name, row_idx):
    """Count how many regions have this service as 'General Availability' for the language"""
    return f'COUNTIFS(Service_Input!$B:$B,"{service_name}",Service_Input!$C:$C,$A{row_idx},Service_Input!$D:$D,"General Availability")'

def is_ga_locally(service_name, row_idx):
    """Check if service is General Availability in selected region"""
    return f'{get_local_status(service_name, row_idx)}="General Availability"'

def is_la_locally(service_name, row_idx):
    """Check if service is Limited Availability in selected region"""
    return f'{get_local_status(service_name, row_idx)}="Limited Availability"'

def is_ga_anywhere(service_name, row_idx):
    """Check if service is GA in at least one region"""
    return f'{get_ga_count_globally(service_name, row_idx)}>=1'

def is_available(service_name, row_idx):
    """Check if service is available (GA locally, LA locally, or GA via cross-region)"""
    # Available = GA locally OR LA locally OR (Not supported locally but GA somewhere)
    return f'OR({is_ga_locally(service_name, row_idx)},{is_la_locally(service_name, row_idx)},{is_ga_anywhere(service_name, row_idx)})'

# Product dependencies
product_deps = {
    "SSA": ["ASR", "SSA-LLM", "KaaS"],
    "RTGA": ["ASR", "Intent (UniFit)", "Entity (NER/Spacy)", "Gen AI Summary", "Gen AI Disposition", "Redaction", "KaaS", "Pro-active Service"],
    "CIA": ["ASR", "Gen AI Disposition", "Redaction", "Conversation Facts"],
    "CRA": ["ASR", "Redaction"]
}

# Rows (Languages) & DEPENDENCY LOGIC
for i, lang in enumerate(languages):
    r = start_row + 1 + i
    ws_prod.cell(row=r, column=1, value=lang).font = Font(bold=True)

    for prod_idx, (product, deps) in enumerate(product_deps.items()):
        col = start_col + prod_idx
        total_deps = len(deps)

        # Count GA locally (for Full Support In-Region check)
        ga_locally_conditions = [f'IF({is_ga_locally(dep, r)},1,0)' for dep in deps]
        count_ga_locally = f'({"+".join(ga_locally_conditions)})'

        # Count GA possible (locally or cross-region) for Full Support Cross-Region check
        ga_possible_conditions = [f'IF({is_ga_anywhere(dep, r)},1,0)' for dep in deps]
        count_ga_possible = f'({"+".join(ga_possible_conditions)})'

        # Count available (GA local, LA local, or GA cross-region) for availability %
        available_conditions = [f'IF({is_available(dep, r)},1,0)' for dep in deps]
        count_available = f'({"+".join(available_conditions)})'

        # Availability percentage
        availability_pct = f'({count_available}/{total_deps})'

        # Build the formula
        # Priority:
        # 1. 100% GA locally → Full Support (In-Region)
        # 2. 100% GA possible (locally or cross-region) → Full Support (Cross-Region)
        # 3. Availability >= 70% → Limited Availability
        # 4. Availability < 70% → Not Supported
        formula = (
            f'=IF({count_ga_locally}={total_deps}, "Full Support (In-Region)", '
            f'IF({count_ga_possible}={total_deps}, "Full Support (Cross-Region)", '
            f'IF({availability_pct}>=0.7, "Limited Availability", '
            f'"Not Supported")))'
        )

        cell = ws_prod.cell(row=r, column=col, value=formula)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

apply_product_formatting(ws_prod, start_row, start_col, len(products), len(languages))
ws_prod.column_dimensions['A'].width = 15
ws_serv.column_dimensions['A'].width = 15

# Save
wb.save(filename)
print(f"File '{filename}' created successfully.")
